import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import ttk
import tkinter.messagebox as messagebox
from tkinter import simpledialog
from datetime import datetime


# 创建主窗口
root = tk.Tk()
root.title("手机出入库管理")


# 固定的 Excel 文件名称
excel_file_name = "手机出入库.xlsx"


# 全局变量，用于存储 Excel 工作簿和工作表对象
workbook = None
worksheet = None


# 存储 result_table 项和 Excel 行号的映射关系
item_row_mapping = {}


# 尝试加载固定的 Excel 文件，如果不存在则创建新的
try:
    workbook = load_workbook(excel_file_name)
    worksheet = workbook.active
    print(f"已成功加载文件：{excel_file_name}")
except FileNotFoundError:
    print(f"{excel_file_name} 文件不存在，创建新文件。")
    workbook = Workbook()
    worksheet = workbook.active
    workbook.save(excel_file_name)


# 写入数据到 Excel 的函数（按照新的字段顺序传入数据）
def write_to_excel():
    global workbook, worksheet
    in_date = in_date_entry.get()
    brand = brand_entry.get()
    model = model_entry.get()
    runtime_memory = runtime_memory_entry.get()
    storage_memory = storage_memory_entry.get()
    color = color_entry.get()
    serial_number = serial_number_entry.get()
    purchase_price = purchase_price_entry.get()
    selling_price = selling_price_entry.get()
    remark = remark_entry.get()
    sell_date = sell_date_entry.get()


    # 如果售价为空，将其设为 0
    if not selling_price:
        selling_price = "0"


    # 将品牌和型号转换为大写形式
    brand = brand.upper()
    model = model.upper()


    # 组合内存信息
    memory = f"{runtime_memory}+{storage_memory}"


    new_row = [in_date, brand, model, memory, color, serial_number, purchase_price, sell_date, selling_price, remark]
    worksheet.append(new_row)
    workbook.save(excel_file_name)
    status_label.config(text="数据已写入 Excel 文件")
    query_data()  # 刷新查询结果
    update_total()  # 写入数据后更新总和


# 更新单元格内容的函数（用于更新进价、售价、备注的值）
def update_cell(new_value, item, col):
    try:
        current_values = list(result_table.item(item, "values"))
        current_values[col] = new_value
        result_table.item(item, values=current_values)


        # 更新对应的 Excel 工作表中的数据
        global workbook, worksheet
        row_index = item_row_mapping[item]  # 使用映射关系找到正确的行号
        cell_to_update = worksheet.cell(row=row_index, column=col + 1)
        cell_to_update.value = new_value
        workbook.save(excel_file_name)
    except ValueError:
        messagebox.showerror("错误", f"请输入有效的{['进价', '售价', '备注'][col - 6]}数值！")


# 查询数据的函数
def query_data():
    global workbook, worksheet
    query_type = query_type_var.get()
    query_value = query_entry.get()
    result_table.delete(*result_table.get_children())
    global item_row_mapping  # 声明全局变量
    item_row_mapping = {}  # 清空映射关系


    # 如果查询值为空，显示全部数据
    if not query_value:
        for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            item = result_table.insert('', 'end', values=row)
            item_row_mapping[item] = index  # 存储映射关系
        return


    found = False
    query_value = query_value.upper()  # 将查询值转换为大写
    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if query_type == "串号":
            serial_number_str = str(row[5])
            if len(serial_number_str) == 15 and len(query_value) == 15:
                same_digit_count = 0
                for i in range(15):
                    if serial_number_str[i] == query_value[i]:
                        same_digit_count += 1
                if same_digit_count >= 12:
                    item = result_table.insert('', 'end', values=row)
                    item_row_mapping[item] = index  # 存储映射关系
                    found = True
        elif query_type == "品牌" and str(row[1]).upper() == query_value:
            item = result_table.insert('', 'end', values=row)
            item_row_mapping[item] = index  # 存储映射关系
            found = True
        elif query_type == "型号" and str(row[2]).upper() == query_value:
            item = result_table.insert('', 'end', values=row)
            item_row_mapping[item] = index  # 存储映射关系
            found = True
        elif query_type == "出售日期":
            # 将用户输入的日期字符串转换为日期对象
            try:
                query_date = datetime.strptime(query_value, "%Y.%m.%d")
                excel_date_str = str(row[7])  # 出售日期所在列的索引为 7
                excel_date = datetime.strptime(excel_date_str, "%Y.%m.%d")
                if excel_date == query_date:
                    item = result_table.insert('', 'end', values=row)
                    item_row_mapping[item] = index  # 存储映射关系
                    found = True
            except ValueError:
                pass
    if not found:
        result_table.insert('', 'end', values=["未找到对应数据"])


# 编辑进价的函数
def edit_purchase_price():
    selected_item = result_table.selection()
    if selected_item:
        selected_item = selected_item[0]
        current_values = list(result_table.item(selected_item, "values"))
        current_price = current_values[6]
        new_price = simpledialog.askfloat("输入进价", "请输入新的进价：", minvalue=0, initialvalue=current_price)
        if new_price is not None:
            update_cell("{:.1f}".format(new_price), selected_item, 6)
            update_total()  # 编辑进价后更新总和


# 编辑售价的函数
def edit_selling_price():
    selected_item = result_table.selection()
    if selected_item:
        selected_item = selected_item[0]
        current_values = list(result_table.item(selected_item, "values"))
        current_price = current_values[8]
        new_price = simpledialog.askfloat("输入售价", "请输入新的售价：", minvalue=0, initialvalue=current_price)
        if new_price is not None:
            update_cell("{:.1f}".format(new_price), selected_item, 8)
            update_total()  # 编辑售价后更新总和


# 编辑备注的函数
def edit_remark():
    selected_item = result_table.selection()
    if selected_item:
        selected_item = selected_item[0]
        current_values = list(result_table.item(selected_item, "values"))
        current_remark = current_values[9]
        new_remark = simpledialog.askstring("输入备注", "请输入新的备注：", initialvalue=current_remark)
        if new_remark is not None:
            update_cell(new_remark, selected_item, 9)


# 编辑串号的函数
def edit_serial_number():
    selected_item = result_table.selection()
    if selected_item:
        selected_item = selected_item[0]
        current_values = list(result_table.item(selected_item, "values"))
        current_serial_number = current_values[5]
        new_serial_number = simpledialog.askstring("输入串号", "请输入新的串号：", initialvalue=current_serial_number)
        if new_serial_number is not None:
            update_cell(new_serial_number, selected_item, 5)


# 编辑出售日期的函数
def edit_sell_date():
    selected_item = result_table.selection()
    if selected_item:
        selected_item = selected_item[0]
        current_values = list(result_table.item(selected_item, "values"))
        current_sell_date = current_values[7]
        new_sell_date = simpledialog.askstring("输入出售日期", "请输入新的出售日期：", initialvalue=current_sell_date)
        if new_sell_date is not None:
            update_cell(new_sell_date, selected_item, 7)
            update_total()


# 删除数据的函数
def delete_data():
    selected_item = result_table.selection()
    if selected_item:
        row_index = item_row_mapping[selected_item[0]]  # 使用映射关系找到正确的行号
        worksheet.delete_rows(row_index)
        workbook.save(excel_file_name)
        result_table.delete(selected_item)
        update_total()  # 删除数据后更新总和


# 计算进价和售价总和的函数
def calculate_total():
    global workbook, worksheet
    total_purchase_price = 0
    total_selling_price = 0
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 7:  # 确保行中包含进价数据
            try:
                if row[6] is not None:  # 检查进价是否为 None
                    total_purchase_price += float(row[6])
            except ValueError:
                pass
        if len(row) >= 9:  # 确保行中包含售价数据
            # 当售价为 None 时默认为 0
            selling_price = 0 if row[8] is None else row[8]
            try:
                total_selling_price += float(selling_price)
            except ValueError:
                pass
    return total_purchase_price, total_selling_price


# 更新显示的总和
def update_total():
    total_purchase_price, total_selling_price = calculate_total()
    total_label.config(text=f"进价总和: {total_purchase_price:.1f}, 售价总和: {total_selling_price:.1f}")


# 创建左侧框架用于写入部分
left_frame = tk.Frame(root)
left_frame.pack(side=tk.LEFT, padx=10, pady=10)


in_date_label = tk.Label(left_frame, text="入库日期：", font=("SimSun", 14))
in_date_label.grid(row=0, column=0, sticky=tk.W, pady=2)
in_date_entry = tk.Entry(left_frame, font=("SimSun", 14))
in_date_entry.grid(row=0, column=1, pady=2)


brand_label = tk.Label(left_frame, text="品牌：", font=("SimSun", 14))
brand_label.grid(row=1, column=0, sticky=tk.W, pady=2)
brand_entry = tk.Entry(left_frame, font=("SimSun", 14))
brand_entry.grid(row=1, column=1, pady=2)


model_label = tk.Label(left_frame, text="型号：", font=("SimSun", 14))
model_label.grid(row=2, column=0, sticky=tk.W, pady=2)
model_entry = tk.Entry(left_frame, font=("SimSun", 14))
model_entry.grid(row=2, column=1, pady=2)


memory_label = tk.Label(left_frame, text="内存：", font=("SimSun", 14))
memory_label.grid(row=3, column=0, sticky=tk.W, pady=2)


# 创建一个框架来放置两个内存输入框
memory_frame = tk.Frame(left_frame)
memory_frame.grid(row=3, column=1, pady=2)


# 运行内存输入框
runtime_memory_entry = tk.Entry(memory_frame, font=("SimSun", 14), width=9)  # 适当调小输入框宽度
runtime_memory_entry.pack(side=tk.LEFT)


# 加号标签
plus_label = tk.Label(memory_frame, text="+", font=("SimSun", 14))
plus_label.pack(side=tk.LEFT)


# 固态内存输入框
storage_memory_entry = tk.Entry(memory_frame, font=("SimSun", 14), width=9)  # 适当调小输入框宽度
storage_memory_entry.pack(side=tk.LEFT)


color_label = tk.Label(left_frame, text="颜色：", font=("SimSun", 14))
color_label.grid(row=4, column=0, sticky=tk.W, pady=2)
color_entry = tk.Entry(left_frame, font=("SimSun", 14))
color_entry.grid(row=4, column=1, pady=2)


serial_number_label = tk.Label(left_frame, text="串号：", font=("SimSun", 14))
serial_number_label.grid(row=5, column=0, sticky=tk.W, pady=2)
serial_number_entry = tk.Entry(left_frame, font=("SimSun", 14))
serial_number_entry.grid(row=5, column=1, pady=2)


purchase_price_label = tk.Label(left_frame, text="进价：", font=("SimSun", 14))
purchase_price_label.grid(row=6, column=0, sticky=tk.W, pady=2)
purchase_price_entry = tk.Entry(left_frame, font=("SimSun", 14))
purchase_price_entry.grid(row=6, column=1, pady=2)


sell_date_label = tk.Label(left_frame, text="出售日期：", font=("SimSun", 14))
sell_date_label.grid(row=7, column=0, sticky=tk.W, pady=2)
sell_date_entry = tk.Entry(left_frame, font=("SimSun", 14))
sell_date_entry.grid(row=7, column=1, pady=2)


selling_price_label = tk.Label(left_frame, text="售价：", font=("SimSun", 14))
selling_price_label.grid(row=8, column=0, sticky=tk.W, pady=2)
selling_price_entry = tk.Entry(left_frame, font=("SimSun", 14))
selling_price_entry.grid(row=8, column=1, pady=2)


remark_label = tk.Label(left_frame, text="备注：", font=("SimSun", 14))
remark_label.grid(row=9, column=0, sticky=tk.W, pady=2)
remark_entry = tk.Entry(left_frame, font=("SimSun", 14))
remark_entry.grid(row=9, column=1, pady=2)


write_button = tk.Button(left_frame, text="写入 Excel", font=("SimSun", 14), command=write_to_excel)
write_button.grid(row=10, columnspan=2, pady=5)


status_label = tk.Label(left_frame, text="", font=("SimSun", 14))
status_label.grid(row=11, columnspan=2)


# 创建右侧框架用于查询结果部分
right_frame = tk.Frame(root)
right_frame.pack(side=tk.RIGHT, padx=10, pady=10)


query_type_label = tk.Label(right_frame, text="查询依据：", font=("SimSun", 18))
query_type_label.pack(pady=5)


query_type_var = tk.StringVar()
query_type_menu = tk.OptionMenu(right_frame, query_type_var, "串号", "品牌", "型号", "出售日期")
query_type_menu['menu'].config(font=("SimSun", 16))
query_type_menu.pack(pady=5)


query_label = tk.Label(right_frame, text="查询值：", font=("SimSun", 16))
query_label.pack(pady=5)
query_entry = tk.Entry(right_frame, font=("SimSun", 16))
query_entry.pack(pady=5)


query_button = tk.Button(right_frame, text="查询", font=("SimSun", 16), command=query_data)
query_button.pack(pady=10)


# 用于放置编辑按钮的框架
edit_button_frame = tk.Frame(right_frame)
edit_button_frame.pack(pady=5)


# 编辑进价按钮
edit_purchase_price_button = tk.Button(edit_button_frame, text="编辑进价", font=("SimSun", 16), command=edit_purchase_price)
edit_purchase_price_button.pack(side=tk.LEFT, padx=5)


# 编辑售价按钮
edit_selling_price_button = tk.Button(edit_button_frame, text="编辑售价", font=("SimSun", 16), command=edit_selling_price)
edit_selling_price_button.pack(side=tk.LEFT, padx=5)


# 编辑备注按钮
edit_remark_button = tk.Button(edit_button_frame, text="编辑备注", font=("SimSun", 16), command=edit_remark)
edit_remark_button.pack(side=tk.LEFT, padx=5)


# 编辑串号按钮
edit_serial_number_button = tk.Button(edit_button_frame, text="编辑串号", font=("SimSun", 16), command=edit_serial_number)
edit_serial_number_button.pack(side=tk.LEFT, padx=5)


# 编辑出售日期按钮
edit_sell_date_button = tk.Button(edit_button_frame, text="编辑出售日期", font=("SimSun", 16), command=edit_sell_date)
edit_sell_date_button.pack(side=tk.LEFT, padx=5)


# 删除数据按钮
delete_button = tk.Button(right_frame, text="删除数据", font=("SimSun", 16), command=delete_data)
delete_button.pack(pady=5)


# 创建表格展示查询结果
style = ttk.Style()
style.configure("Treeview", font=('SimSun', 14))  # 设置字体大小为 14
result_table = ttk.Treeview(right_frame, columns=("入库日期", "品牌", "型号", "内存", "颜色", "串号", "进价", "出售日期", "售价", "备注"),
                        show="headings")
result_table.column("入库日期", width=100)
result_table.column("品牌", width=60)
result_table.column("型号", width=120)
result_table.column("内存", width=70)
result_table.column("颜色", width=100)
result_table.column("串号", width=180)
result_table.column("进价", width=80)
result_table.column("出售日期", width=100)
result_table.column("售价", width=80)
result_table.column("备注", width=60)
for col in ("入库日期", "品牌", "型号", "内存", "颜色", "串号", "进价", "出售日期", "售价", "备注"):
    result_table.heading(col, text=col)
result_table.pack(pady=10)


# 显示总和的标签
total_label = tk.Label(right_frame, text="", font=("SimSun", 16))
total_label.pack(pady=5)


# 初始化时查询所有数据
query_data()


# 计算并显示总和
update_total()


# 运行主循环
root.mainloop()